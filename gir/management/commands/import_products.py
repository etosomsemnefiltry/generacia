from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook
import os
from gir.models import Category, Product


class Command(BaseCommand):
    help = 'Імпорт товарів з Excel файлу "Кошторис Генерація_2.0 (1).xlsx" з усіх вкладок'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='FILES/Кошторис Генерація_2.0 (1).xlsx',
            help='Шлях до Excel файлу'
        )
        parser.add_argument(
            '--sheets',
            nargs='+',
            default=None,
            help='Список вкладок для обробки (якщо не вказано - обробляємо всі)'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        specific_sheets = options['sheets']
        
        if not os.path.exists(file_path):
            self.stdout.write(
                self.style.ERROR(f'Файл не знайдено: {file_path}')
            )
            return
        
        try:
            # Завантажуємо Excel
            self.stdout.write(f'Завантажую файл: {file_path}')
            workbook = load_workbook(file_path, data_only=True)
            
            # Визначаємо які вкладки обробляти
            if specific_sheets:
                sheets_to_process = [s for s in specific_sheets if s in workbook.sheetnames]
                if not sheets_to_process:
                    self.stdout.write(
                        self.style.ERROR(f'Жодної з вказаних вкладок не знайдено: {specific_sheets}')
                    )
                    return
            else:
                # Обробляємо всі вкладки, крім системних
                sheets_to_process = [s for s in workbook.sheetnames if not self._is_system_sheet(s)]
            
            self.stdout.write(f'Вкладки для обробки: {sheets_to_process}')
            
            total_created = 0
            total_categories = 0
            
            # Обробляємо кожну вкладку
            for sheet_name in sheets_to_process:
                self.stdout.write(f'\n{"="*50}')
                self.stdout.write(f'Обробляю вкладку: {sheet_name}')
                self.stdout.write(f'{"="*50}')
                
                sheet_created, sheet_categories = self._process_sheet(workbook[sheet_name], sheet_name)
                total_created += sheet_created
                total_categories += sheet_categories
            
            self.stdout.write(f'\n{"="*50}')
            self.stdout.write(
                self.style.SUCCESS(
                    f'Імпорт завершено! Загалом створено: {total_created} товарів, {total_categories} категорій'
                )
            )
            self.stdout.write(f'{"="*50}')
            
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Помилка імпорту: {e}')
            )
    
    def _is_system_sheet(self, sheet_name):
        """Перевіряє чи є вкладка системною"""
        system_sheets = [
            'Зміст', 'Титул', 'Зміст_2', 'Титул_2',
            'Кошторис', 'Кошторис_2', 'Кошторис_3',
            'Відомість', 'Відомість_2', 'Відомість_3',
            'Сума', 'Сума_2', 'Сума_3',
            'Копія аркуша', 'ТЕХ_ЛИСТ', 'ТЗ', 'Розрахунок',
            'Вартість монтажу', 'Показники', 'Довідкові дані',
            'Інструкції', 'РЕЗЕРВ аркуша', 'Прайс'
        ]
        return any(sys_name.lower() in sheet_name.lower() for sys_name in system_sheets)
    
    def _process_sheet(self, worksheet, sheet_name):
        """Обробляє одну вкладку Excel"""
        created_count = 0
        categories_count = 0
        
        try:
            # Знаходимо заголовки
            headers = self._find_headers(worksheet)
            if not headers:
                self.stdout.write(f'Заголовки не знайдено у вкладці {sheet_name}')
                return 0, 0
            
            self.stdout.write(f'Знайдені заголовки: {headers}')
            
            # Маппінг полів Excel → Django
            field_mapping = {
                'Категорія': 'category',
                'Найменування': 'title',
                'Од.вим': 'unit',
                'Облікова ціна б.г. з ПДВ $': 'cost_price',
                'Гуртова ціна з ПДВ, авт% $': 'wholesale_price',
                'Артикул': 'sku',
                'Кількість модулів': 'modules_count',
                'Вільно на складі': 'stock_quantity',
                'Посилання': 'external_link'
            }
            
            # Знаходимо індекси колонок
            column_indices = {}
            for i, header in enumerate(headers):
                for excel_field, django_field in field_mapping.items():
                    if excel_field in str(header):
                        column_indices[django_field] = i
                        break
            
            self.stdout.write(f'Маппінг колонок: {column_indices}')
            
            # Перевіряємо чи є обов'язкові поля
            if 'category' not in column_indices or 'title' not in column_indices:
                self.stdout.write(f'У вкладці {sheet_name} відсутні обов\'язкові поля (Категорія або Найменування)')
                return 0, 0
            
            # Імпортуємо дані
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                try:
                    # Отримуємо значення з рядка
                    row_data = {}
                    for field, col_idx in column_indices.items():
                        if col_idx < len(row):
                            value = row[col_idx].value
                            row_data[field] = value
                    
                    # Перевіряємо обов'язкові поля
                    if not row_data.get('title') or not row_data.get('category'):
                        continue
                    
                    # Створюємо або отримуємо категорію
                    category_name = str(row_data['category']).strip()
                    if not category_name:
                        continue
                        
                    category, created = Category.objects.get_or_create(
                        name=category_name,
                        defaults={
                            'slug': self._generate_slug(category_name),
                            'is_active': True
                        }
                    )
                    
                    if created:
                        categories_count += 1
                        self.stdout.write(f'Створено категорію: {category_name}')
                    
                    # Перевіряємо чи існує товар
                    product_title = str(row_data['title']).strip()
                    product_sku = str(row_data.get('sku', '')).strip()
                    
                    # Шукаємо по артикулу або назві
                    existing_product = None
                    if product_sku:
                        existing_product = Product.objects.filter(sku=product_sku).first()
                    
                    if not existing_product and product_title:
                        existing_product = Product.objects.filter(title=product_title).first()
                    
                    # Підготовляємо дані для товара
                    product_data = {
                        'category': category,
                        'title': product_title,
                        'unit': str(row_data.get('unit', '')).strip(),
                        'sku': product_sku,
                        'modules_count': self._parse_int(row_data.get('modules_count')),
                        'stock_quantity': self._parse_int(row_data.get('stock_quantity'), default=0),
                        'external_link': str(row_data.get('external_link', '')).strip(),
                        'is_active': True
                    }
                    
                    # Обробляємо ціни
                    if row_data.get('cost_price'):
                        product_data['cost_price'] = self._parse_decimal(row_data['cost_price'])
                    
                    if row_data.get('wholesale_price'):
                        product_data['wholesale_price'] = self._parse_decimal(row_data['wholesale_price'])
                    
                    if existing_product:
                        # Оновлюємо існуючий товар
                        for field, value in product_data.items():
                            if value is not None and value != '':
                                setattr(existing_product, field, value)
                        existing_product.save()
                        self.stdout.write(f'Оновлено товар: {product_title}')
                    else:
                        # Створюємо новий товар
                        product = Product.objects.create(**product_data)
                        created_count += 1
                        self.stdout.write(f'Створено товар: {product_title}')
                    
                    if row_num % 100 == 0:
                        self.stdout.write(f'Оброблено {row_num} рядків...')
                        
                except Exception as e:
                    self.stdout.write(
                        self.style.WARNING(f'Помилка в рядку {row_num}: {e}')
                    )
                    continue
            
            self.stdout.write(
                self.style.SUCCESS(
                    f'Вкладка {sheet_name}: створено {created_count} товарів, {categories_count} категорій'
                )
            )
            
            return created_count, categories_count
            
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Помилка обробки вкладки {sheet_name}: {e}')
            )
            return 0, 0
    
    def _find_headers(self, worksheet):
        """Знаходить заголовки у вкладці"""
        headers = []
        for row in worksheet.iter_rows(min_row=1, max_row=20):  # Шукаємо в перших 20 рядках
            for cell in row:
                if cell.value:
                    headers.append(cell.value)
                else:
                    headers.append('')
            # Перевіряємо чи є заголовок "Категорія"
            if any('Категорія' in str(h) for h in headers):
                return headers
        return []
    
    def _generate_slug(self, name):
        """Генерує slug з назви категорії"""
        import re
        # Транслітерація українських літер
        translit_map = {
            'а': 'a', 'б': 'b', 'в': 'v', 'г': 'h', 'ґ': 'g', 'д': 'd', 'е': 'e', 'є': 'ye',
            'ж': 'zh', 'з': 'z', 'и': 'y', 'і': 'i', 'ї': 'yi', 'й': 'y', 'к': 'k', 'л': 'l',
            'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
            'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'shch', 'ь': '',
            'ю': 'yu', 'я': 'ya'
        }
        
        # Транслітеруємо
        slug = name.lower()
        for ukr, eng in translit_map.items():
            slug = slug.replace(ukr, eng)
        
        # Замінюємо пробіли та спецсимволи на дефіси
        slug = re.sub(r'[^a-z0-9]+', '-', slug)
        slug = re.sub(r'^-+|-+$', '', slug)  # Прибираємо дефіси з початку та кінця
        
        return slug
    
    def _parse_int(self, value, default=None):
        """Парсинг цілого числа"""
        if value is None:
            return default
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return default
    
    def _parse_decimal(self, value):
        """Парсинг десяткового числа"""
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
